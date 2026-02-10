"""Fill Excel templates with analysis_pack data using openpyxl.

This module writes VALUES only. It does NOT recompute any formulas.
The template's output sheet contains narrative string-concatenation formulas
that reference these cells. Formula recalculation happens in the PDF export
step via win32com (Excel).
"""
from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from .registry import (
    TRANSFORMS,
    _resolve,
    get_template_config,
)


TEMPLATE_DIR = Path(__file__).resolve().parent.parent.parent / "templates" / "excel"


def _resolve_source(
    source: str,
    analysis_pack: dict,
    meta: dict,
    computed: dict,
) -> Any:
    """Resolve a source path to its value."""
    if source.startswith("meta."):
        key = source[5:]
        return meta.get(key)
    if source.startswith("_computed."):
        key = source[10:]
        return computed.get(key)
    if source.startswith("legs."):
        parts = source.split(".")
        try:
            idx = int(parts[1])
            legs = analysis_pack.get("legs", [])
            if idx < len(legs):
                leg = legs[idx]
                field = parts[2] if len(parts) > 2 else None
                return leg.get(field) if field else leg
        except (ValueError, IndexError):
            pass
        return None
    return _resolve(analysis_pack, source)


def _parse_cell_ref(cell_ref: str) -> tuple[str, str]:
    """Parse 'sheet!A1' -> ('sheet', 'A1'). If no sheet prefix, returns ('', cell)."""
    if "!" in cell_ref:
        sheet, cell = cell_ref.split("!", 1)
        return sheet, cell
    return "", cell_ref


def fill_template(
    template_key: str,
    analysis_pack: dict,
    meta: dict,
    output_path: Optional[str] = None,
) -> str:
    """Fill an Excel template with analysis_pack data.

    Args:
        template_key: Registry key (e.g. "template_covered_call")
        analysis_pack: Output from build_analysis_pack()
        meta: Dict with keys like underlying_ticker, option_ticker_leg1,
              shares, cio_rating, etc.
        output_path: Where to save filled xlsx. If None, saves to out/ dir.

    Returns:
        Path to the filled xlsx file.
    """
    config = get_template_config(template_key)
    if config is None:
        raise ValueError(f"Unknown template: {template_key}")

    template_file = TEMPLATE_DIR / config["file"]
    if not template_file.exists():
        raise FileNotFoundError(
            f"Template file not found: {template_file}\n"
            f"Place the template in: {TEMPLATE_DIR}"
        )

    if output_path is None:
        out_dir = Path("out")
        out_dir.mkdir(exist_ok=True)
        stem = template_file.stem
        output_path = str(out_dir / f"{stem}_filled.xlsx")

    shutil.copy2(str(template_file), output_path)

    # Build computed values
    expiry_str = analysis_pack.get("strategy", {}).get("expiry")
    today = date.today()
    dte = None
    if expiry_str:
        try:
            expiry_date = date.fromisoformat(str(expiry_str)[:10])
            dte = (expiry_date - today).days
        except (ValueError, TypeError):
            pass

    legs = analysis_pack.get("legs", [])
    bid_ask_spread = 0
    if legs:
        leg0 = legs[0]
        bid = leg0.get("bid")
        ask = leg0.get("ask")
        if bid is not None and ask is not None:
            try:
                bid_ask_spread = float(ask) - float(bid)
            except (ValueError, TypeError):
                pass

    # Spot minus strike for Short Put (OTM amount display)
    spot = analysis_pack.get("underlying", {}).get("spot") or 0
    strike_0 = 0
    if legs:
        try:
            strike_0 = float(legs[0].get("strike", 0))
        except (ValueError, TypeError):
            pass

    # Bid-ask spread for leg 2 (multi-leg templates)
    bid_ask_spread_1 = 0
    if len(legs) > 1:
        bid1 = legs[1].get("bid")
        ask1 = legs[1].get("ask")
        if bid1 is not None and ask1 is not None:
            try:
                bid_ask_spread_1 = float(ask1) - float(bid1)
            except (ValueError, TypeError):
                pass

    computed = {
        "dte": dte,
        "now": datetime.now(),
        "trade_cost": 0,
        "bid_ask_spread": bid_ask_spread,
        "bid_ask_spread_1": bid_ask_spread_1,
        "spot_minus_strike": spot - strike_0,
    }

    wb = load_workbook(output_path)
    injections = config.get("injections", [])
    errors: list[str] = []

    for inj in injections:
        cell_ref = inj["cell"]
        source = inj["source"]
        transform_name = inj.get("transform")
        default = inj.get("default")
        label = inj.get("label", cell_ref)

        value = _resolve_source(source, analysis_pack, meta, computed)

        if value is None and default is not None:
            value = default

        if transform_name and transform_name in TRANSFORMS:
            try:
                value = TRANSFORMS[transform_name](value)
            except Exception as e:
                errors.append(f"[TRANSFORM] {label} ({cell_ref}): {e}")
                continue

        if value is None:
            continue

        sheet_name, cell_addr = _parse_cell_ref(cell_ref)
        if sheet_name not in wb.sheetnames:
            errors.append(f"[SHEET] {label}: sheet '{sheet_name}' not found")
            continue

        try:
            ws = wb[sheet_name]
            ws[cell_addr] = value
        except Exception as e:
            errors.append(f"[WRITE] {label} ({cell_ref}): {e}")

    if errors:
        print(f"[TEMPLATE_FILL] {len(errors)} warnings:")
        for err in errors:
            print(f"  {err}")

    wb.save(output_path)
    wb.close()

    print(f"[TEMPLATE_FILL] Saved filled template: {output_path}")
    return output_path
